from locust import HttpUser, task, TaskSet, SequentialTaskSet, constant, between
import time
import openpyxl


class MyReq(SequentialTaskSet):
    def on_start(self):
        self.login()

    @task
    def login(self):
        start = time.time()
        response1 = self.client.get('/')
        print("get method status is", response1.status_code)
        csrf_token = response1.cookies['csrftoken']
        response2 = self.client.post('/', {'username': 'admin', 'password': 'adm12'},
                                         headers={'X-CSRFToken': csrf_token})
        end = time.time()
        print("post method status is", response2.status_code)
        print(f"DEBUG: login response.status_code = {response2.status_code}")
        # after logging status code
        print("starting time ", start)
        print("ending time ", end)
        print("total time= ", (end - start))
        print(" ")

    def data(self):
        wb = openpyxl.load_workbook("scenario2ID1111.xlsx")
        sh = wb.active

        # iterate through excel and display data
        for i in range(1, sh.max_row + 1):
            print("\n")
            print("Row ", i, " data :")

            for j in range(1, sh.max_column + 1):
                cell_obj = sh.cell(row=i, column=j)
                print(cell_obj.value, end=" ")
class MySeq(HttpUser):
    wait_time = between(1, 5)
    host = "http://10.91.28.85"
    tasks = [MyReq]
