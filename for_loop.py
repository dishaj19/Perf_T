
import openpyxl

wb = openpyxl.load_workbook("scenario2ID1111.xlsx")

sh = wb.active

# iterate through excel and display data
for i in range(1, sh.max_row + 1):
    print("\n")
    print("Row ", i, " data :")

    for j in range(1, sh.max_column + 1):
        cell_obj = sh.cell(row=i, column=j)
        print(cell_obj.value, end=" ")




